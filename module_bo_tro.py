import numpy as np
import math
import openpyxl
import cv2
import json
import time
import matplotlib.pyplot as plt
import pickle
import random
import os
import pandas as pd
from lib_TU_me import LaneEval

# Define the font properties
font = cv2.FONT_HERSHEY_SIMPLEX
font_scale = 0.4
font_thickness = 1
font_color = (0, 0, 255)  # BGR format (blue, green, red)

wb = openpyxl.Workbook()
ws = wb.active

y_start = 320; y_stop = 450; x_ext_bot = 500; x_ext_top = 200; x_mid = 820-60
h = 590; w = 1640
h_roi = y_stop - y_start; w_roi = x_ext_bot
y_roi = y_stop - y_start; x_roi = x_ext_bot

phi_min = 5; phi_max = 85

Rho_max_ram = round(math.sqrt((y_stop-y_start)**2 + (x_ext_bot)**2)) # idea 3: origin of coordinate in top leftmost corner, in top rightmost corner
phi_max_ram = phi_max

t_g = 30; t_g_edge_max = 50

roi_left = np.array([((w_roi-x_ext_top),0), (w_roi,0), (w_roi,h_roi), (0,h_roi)])
mask_left = np.zeros(shape=(h_roi,w_roi), dtype=np.uint8)
cv2.fillPoly(mask_left, [roi_left], (255))

roi_right = np.array([(0,0), (x_ext_top,0), (w_roi,h_roi), (0,h_roi)])
mask_right = np.zeros(shape=(h_roi,w_roi), dtype=np.uint8)
cv2.fillPoly(mask_right, [roi_right], (255))

roi_left_draw = np.array([((x_mid-x_ext_top),y_start), (x_mid,y_start), (x_mid,y_stop), (x_mid-x_ext_bot,y_stop)])
roi_right_draw = np.array([(x_mid,y_start), (x_mid+x_ext_top,y_start), (x_mid+x_ext_bot,y_stop), (x_mid,y_stop)])

stt = np.arange(w_roi)

class Save_csv():
    def Sobel_get_G(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        sobel_x = cv2.Sobel(img_left, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_left, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=1)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Sobel_get_G_left.csv', img_G_tmp, delimiter=',', fmt='%d')

        sobel_x = cv2.Sobel(img_right, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_right, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=0)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Sobel_get_G_right.csv', img_G_tmp, delimiter=',', fmt='%d')
    def Median5x5_Sobel_get_G(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.medianBlur(img_left, ksize=5)
        img_right = cv2.medianBlur(img_right, ksize=5)

        sobel_x = cv2.Sobel(img_left, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_left, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=1)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Median5x5_Sobel_get_G_left.csv', img_G_tmp, delimiter=',', fmt='%d')

        sobel_x = cv2.Sobel(img_right, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_right, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=0)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Median5x5_Sobel_get_G_right.csv', img_G_tmp, delimiter=',', fmt='%d')
    def Median3x3_Sobel_get_G(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.medianBlur(img_left, ksize=3)
        img_right = cv2.medianBlur(img_right, ksize=3)

        sobel_x = cv2.Sobel(img_left, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_left, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=1)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Median3x3_Sobel_get_G_left.csv', img_G_tmp, delimiter=',', fmt='%d')

        sobel_x = cv2.Sobel(img_right, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_right, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=0)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Median3x3_Sobel_get_G_right.csv', img_G_tmp, delimiter=',', fmt='%d')
    def Gaussian3x3_Sobel_get_G(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.GaussianBlur(img_left, (3, 3), 1.0)
        img_right = cv2.GaussianBlur(img_right, (3, 3), 1.0)

        sobel_x = cv2.Sobel(img_left, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_left, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(img_G.shape[1]-5, 4,-1):
                    if mask_left[i,j] == 0:
                        break
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=1)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
                        #img_G_tmp[i,j] = round(img_G[i,j] + phi / 100., 2)
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Gaussian3x3_Sobel_get_G_left.csv', img_G_tmp, delimiter=',', fmt='%d')

        sobel_x = cv2.Sobel(img_right, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_right, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    if mask_right[i,j] == 0:
                        break
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=0)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
                        #img_G_tmp[i,j] = round(img_G[i,j] + phi / 100., 2)
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Gaussian3x3_Sobel_get_G_right.csv', img_G_tmp, delimiter=',', fmt='%d')
    def Gaussian5x5_Sobel_get_G(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.GaussianBlur(img_left, (5, 5), 1.0)
        img_right = cv2.GaussianBlur(img_right, (5, 5), 1.0)

        sobel_x = cv2.Sobel(img_left, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_left, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=1)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
                        #img_G_tmp[i,j] = round(img_G[i,j] + phi / 100., 2)
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Gaussian5x5_Sobel_get_G_left.csv', img_G_tmp, delimiter=',', fmt='%d')

        sobel_x = cv2.Sobel(img_right, cv2.CV_64F, 1, 0, ksize=3); sobel_y = cv2.Sobel(img_right, cv2.CV_64F, 0, 1, ksize=3)
        img_G = np.round(np.sqrt(sobel_x**2 + sobel_y**2)); img_phi = np.round(np.degrees(np.arctan2(sobel_y, sobel_x)))
        img_G_tmp = np.zeros_like(img_G)
        for i in range(5, img_G.shape[0]-5):
            for j in range(5, img_G.shape[1]-5):
                    phi = img_phi[i,j]; phi = Voting().check_phi(phi,flag_left=0)   
                    if phi > 0:
                        img_G_tmp[i,j] = img_G[i,j]
                        #img_G_tmp[i,j] = round(img_G[i,j] + phi / 100., 2)
        img_G_tmp[0, : ] = stt
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_Gaussian5x5_Sobel_get_G_right.csv', img_G_tmp, delimiter=',', fmt='%d')
    def Grayscale(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = img_left.astype(np.uint16); img_right = img_right.astype(np.uint16)

        img_left[0:5, : ] = 0; img_left[h_roi-6:h_roi, : ] = 0; img_left[:, 0:5] = 0; img_left[:, x_mid-6:x_mid] = 0
        img_right[0:5, : ] = 0; img_right[h_roi-6:h_roi, : ] = 0; img_right[:, 0:5] = 0; img_right[:, x_mid-6:x_mid] = 0
        
        img_left[0, : ] = stt; img_right[0, : ] = stt
        
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_left.csv', img_left, delimiter=',', fmt='%d')
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_right.csv', img_right, delimiter=',', fmt='%d')
    def Grayscale_Median5x5(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.medianBlur(img_left, ksize=5)
        img_right = cv2.medianBlur(img_right, ksize=5)

        img_left = img_left.astype(np.uint16); img_right = img_right.astype(np.uint16)

        img_left[0:5, : ] = 0; img_left[h_roi-6:h_roi, : ] = 0; img_left[:, 0:5] = 0; img_left[:, x_mid-6:x_mid] = 0
        img_right[0:5, : ] = 0; img_right[h_roi-6:h_roi, : ] = 0; img_right[:, 0:5] = 0; img_right[:, x_mid-6:x_mid] = 0
        
        img_left[0, : ] = stt; img_right[0, : ] = stt
        
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_median5x5_left.csv', img_left, delimiter=',', fmt='%d')
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_edian5x5_right.csv', img_right, delimiter=',', fmt='%d')
    def Grayscale_Gaussian3x3(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.GaussianBlur(img_left, (3, 3), 1.0)
        img_right = cv2.GaussianBlur(img_right, (3, 3), 1.0)

        img_left = img_left.astype(np.uint16); img_right = img_right.astype(np.uint16)

        img_left[0:5, : ] = 0; img_left[h_roi-6:h_roi, : ] = 0; img_left[:, 0:5] = 0; img_left[:, x_mid-6:x_mid] = 0
        img_right[0:5, : ] = 0; img_right[h_roi-6:h_roi, : ] = 0; img_right[:, 0:5] = 0; img_right[:, x_mid-6:x_mid] = 0
        
        img_left[0, : ] = stt; img_right[0, : ] = stt
        
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_gaussian3x3_left.csv', img_left, delimiter=',', fmt='%d')
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_gaussian3x3_right.csv', img_right, delimiter=',', fmt='%d')
    def Grayscale_Gaussian5x5(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); data_list = []
        gt = json_gt[number_frame]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

        img = cv2.imread(raw_file_old); img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) 

        img_left = img[y_start:y_stop, (x_mid-w_roi):x_mid]
        img_right = img[y_start:y_stop, x_mid:(x_mid+w_roi)]

        img_left = cv2.GaussianBlur(img_left, (5, 5), 1.0)
        img_right = cv2.GaussianBlur(img_right, (5, 5), 1.0)

        img_left = img_left.astype(np.uint16); img_right = img_right.astype(np.uint16)

        img_left[0:5, : ] = 0; img_left[h_roi-6:h_roi, : ] = 0; img_left[:, 0:5] = 0; img_left[:, x_mid-6:x_mid] = 0
        img_right[0:5, : ] = 0; img_right[h_roi-6:h_roi, : ] = 0; img_right[:, 0:5] = 0; img_right[:, x_mid-6:x_mid] = 0
        
        img_left[0, : ] = stt; img_right[0, : ] = stt
        
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_gaussian5x5_left.csv', img_left, delimiter=',', fmt='%d')
        np.savetxt(f'csv/{name}/{raw_file_new[-8:-4]}_gray_gaussian5x5_right.csv', img_right, delimiter=',', fmt='%d')

class Draw():
    def Draw_Roi_and_Lane_True(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        total_frame = len(json_gt); data_list = []

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; 
            raw_file_old = gt['raw_file_old']; lanes = gt['lanes']
            mid_lane_left = gt['mid_lane_left']; mid_lane_right = gt['mid_lane_right']

            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            img = cv2.imread(raw_file_old)

            cv2.polylines(img, [roi_left_draw], isClosed=True, color=(127,255,0), thickness=2)
            cv2.polylines(img, [roi_right_draw], isClosed=True, color=(255,127,0), thickness=2)

            for lane in lanes:
                for pt in lane:
                    x,y = pt[0], pt[1]
                    cv2.circle(img, (x,y), radius=2, color=(200, 127, 30), thickness=-1)

            for pt in mid_lane_left:
                x,y = pt[0], pt[1]
                cv2.circle(img, (x,y), radius=2, color=(0, 255, 0), thickness=-1)
            for pt in mid_lane_right:
                x,y = pt[0], pt[1]
                cv2.circle(img, (x,y), radius=2, color=(0, 255, 0), thickness=-1)

            cv2.imwrite(f'dataset_phan_loai/{name}/{raw_file_new[-8:]}', img)
            
            print(f"{raw_file_new} / {total_frame}:", time.time() - start_time, "seconds") 
    def Draw_Only_Roi(seft,name, number_frame):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        total_frame = len(json_gt); data_list = []
        
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; 
            raw_file_old = gt['raw_file_old']; lanes = gt['lanes']
            mid_lane_left = gt['mid_lane_left']; mid_lane_right = gt['mid_lane_right']
            #if int(raw_file_new[-8:-4]) != number_frame: continue
            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            img = cv2.imread(raw_file_old)

            cv2.polylines(img, [roi_left_draw], isClosed=True, color=(127,255,0), thickness=2)
            cv2.polylines(img, [roi_right_draw], isClosed=True, color=(255,127,0), thickness=2)

            cv2.imwrite(f'{raw_file_new}', img)
            print(f"{raw_file_new} / {total_frame}:", time.time() - start_time, "seconds") 
    def Connect_frame_into_video_demo(seft, name):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        total_frame = len(json_gt); data_list = []

        image_files = []
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']

            #img = cv2.imread(raw_file_old)
            img = cv2.imread(raw_file_new)
            #img = img[y_start:y_stop,x_mid-x_ext_bot:x_mid+x_ext_bot,:]

            cv2.putText(img, f'Frame: {raw_file_new[-8:-4]}', (20,20), font, font_scale+0.5, (255,255,255), font_thickness)
            image_files.append(img)
            print(f'{raw_file_new}/{total_frame}')

        first_image = image_files[0]
        frame_height, frame_width, _ = first_image.shape
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(f'dataset_phan_loai/video/{name}_Frame[{number_frame[0]},{number_frame[1]}]_result.mp4', fourcc, 5, (frame_width, frame_height))
        for image_file in image_files:
            out.write(image_file)
        out.release()
    def Draw_predict_gt_text(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f"json/2lane/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}].json")][0]
        json_gt_02 = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        gts_02 = {l['raw_file_new']: l for l in json_gt_02}
        total_frame = len(json_gt); data_list = []

        # json_no_lane_left = [json.loads(line) for line in open(f'json/true/{name}_no_lane_left_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]
        # json_no_lane_right = [json.loads(line) for line in open(f'json/true/{name}_no_lane_right_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]

        # json_no_lane_left = {l['raw_file_new']: l for l in json_no_lane_left}
        # json_no_lane_right = {l['raw_file_new']: l for l in json_no_lane_right}

        image_files = []
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']
            pred_lane_left = gt['pred_lane_left']; pred_lane_right = gt['pred_lane_right']; 

            gt_02 = gts_02[raw_file_new]
            mid_lane_left = gt_02['mid_lane_left']; mid_lane_right = gt_02['mid_lane_right']; 
            multi_lane_left = gt_02['multi_lane_left']; multi_lane_right = gt_02['multi_lane_right']; 
            angle_right = gt_02['angle_mid_lane_right']; angle_left = gt_02['angle_mid_lane_left']
            lanes = gt_02['lanes']
            
            #if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            # if raw_file_new in json_no_lane_left or raw_file_new in json_no_lane_right:
            #     print(f'{raw_file_new} no lane')
            #     continue
            # if acc_left > 50 and acc_right > 50: continue
            img = cv2.imread(raw_file_old)

            cv2.polylines(img, [roi_left_draw], isClosed=True, color=(127,255,0), thickness=2)
            cv2.polylines(img, [roi_right_draw], isClosed=True, color=(255,127,0), thickness=2)
            cv2.line(img, (x_mid-150, y_start), (x_mid-400, y_stop), (255,127,255), 1)
            cv2.line(img, (x_mid+150, y_start), (x_mid+400, y_stop), (255,127,255), 1)

            angles_left = LaneEval.get_angle(mid_lane_left)
            angles_right = LaneEval.get_angle(mid_lane_right)

            left_thresh = LaneEval.pixel_thresh / np.cos(angles_left)
            right_thresh = LaneEval.pixel_thresh / np.cos(angles_right)

            for l in lanes:
                for pt in l:
                    x,y = pt[0], pt[1]
                    cv2.circle(img, (x,y), radius=1, color=(255, 127, ), thickness=1)
                    #cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)

            for pt in mid_lane_left:
                x,y = pt[0], pt[1]
                cv2.circle(img, (x,y), radius=1, color=(0, 255, ), thickness=1)
                cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)
            for pt in mid_lane_right:
                x,y = pt[0], pt[1]
                cv2.circle(img, (x,y), radius=1, color=(0, 255, ), thickness=1)
                cv2.line(img, (round(x-right_thresh), y), (round(x+right_thresh), y), (127,127,0), 1)

            for pt in pred_lane_left:
                cv2.circle(img, pt, radius=1, color=(0, 0, 255), thickness=1)
            for pt in pred_lane_right:
                cv2.circle(img, pt, radius=1, color=(0, 0, 255), thickness=1)

            # x, y, w, h = (460-20)-10, (y_start-50)-10, 700, (50-10)
            # mask = np.zeros_like(img)
            # cv2.rectangle(mask, (x, y), (x + w, y + h), (255, 255, 255), -1)
            #img = np.where(mask == 255, (255, 255, 255), img)

            # str_tmp = f't_g:{t_g}; pixel_thresh_width_2dege:{pixel_thresh_width_2dege}; denta_gray_max:{denta_gray_max}; ex_bot:{x_ext_bot}; ex_top:{x_ext_top}'
            # cv2.putText(img, str_tmp, (460,y_start-50), font, font_scale, font_color, font_thickness)
            # if raw_file_new in json_no_lane_left:
            #     cv2.putText(img, f'NO LANE LEFT', (460,y_start-50), font, font_scale+0.3, font_color, font_thickness+1)
            # if raw_file_new in json_no_lane_right:
            #     cv2.putText(img, f'NO LANE RIGHT', (x_mid ,y_start-50), font, font_scale+0.3, font_color, font_thickness+1)
            str_tmp = f'Frame: {raw_file_new[-8:-4]} {ratio_left}- {round(acc_left, 2)}%  right: {ratio_right} - {round(acc_right, 2)}%'
            cv2.putText(img, str_tmp, (460-50,y_start-40), font, font_scale+0.3, font_color, font_thickness+1)
            str_tmp = f'angle_left: {round(abs(angle_left),2)} - angle_right: {round(angle_right,2)}'
            cv2.putText(img, str_tmp, (460-50,y_start-20), font, font_scale+0.1, font_color, font_thickness)

            img = img[y_start-70:y_stop+5, x_mid-x_ext_bot-20:x_mid+x_ext_bot+20, ]
            img = cv2.resize(img, (img.shape[1]*4, img.shape[0]*4))
            cv2.imwrite(f'dataset_phan_loai/{name}/{raw_file_new[-8:]}', img)
            image_files.append(img)
            print(f'{raw_file_new}')

        # first_image = image_files[0]
        # frame_height, frame_width, _ = first_image.shape
        # fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        # out = cv2.VideoWriter(f'dataset_phan_loai/video/{name}_wrong_acc_less_than_50.mp4', fourcc, 5, (frame_width, frame_height))
        # for image_file in image_files:
        #     out.write(image_file)
        # out.release()
    def Draw_frame_censorship(seft, name, number_frame, option):
        df = pd.read_excel(f'Analysis/Frame_NO_LINE/{name}_ratio_acc_Frame[{number_frame[0]},{number_frame[1]}].xlsx')
        filtered_rows = df[df['L.Khong co lane'] == f'{option}']
        raw_file_new = filtered_rows['raw_file_new'].tolist()
        raw_file_old = filtered_rows['raw_file_old'].tolist()

        data_list = []
        for i in range(len(raw_file_new)):
            data = {
                "raw_file_new": raw_file_new[i],
                "raw_file_old": raw_file_old[i]
            }
            data_list.append(data)
            print(f'{raw_file_new[i]}')

        data_list = {l['raw_file_new']: l for l in data_list}

        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        total_frame = len(json_gt)

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; raw_file_new = gt['raw_file_new']

            if raw_file_new in data_list:
                img = cv2.imread(f'dataset_phan_loai/{name}/{raw_file_new[-8:]}')
                cv2.imwrite(f'my_frame_processing/censorship/{name}/{option}_left/{raw_file_new[-8:-4]}_{option}_left.jpg', img)
                print(f'{raw_file_new}')

        filtered_rows = df[df['R.Khong co lane'] == f'{option}']
        raw_file_new = filtered_rows['raw_file_new'].tolist()
        raw_file_old = filtered_rows['raw_file_old'].tolist()

        data_list = []
        for i in range(len(raw_file_new)):
            data = {
                "raw_file_new": raw_file_new[i],
                "raw_file_old": raw_file_old[i]
            }
            data_list.append(data)
            print(f'{raw_file_new[i]}')

        data_list = {l['raw_file_new']: l for l in data_list}

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; raw_file_new = gt['raw_file_new']

            if raw_file_new in data_list:
                img = cv2.imread(f'dataset_phan_loai/{name}/{raw_file_new[-8:]}')
                cv2.imwrite(f'my_frame_processing/censorship/{name}/{option}_right/{raw_file_new[-8:-4]}_{option}_right.jpg', img)
                print(f'{raw_file_new}')
class Voting():
    def check_phi(seft,phi,flag_left):
        if flag_left:
            if (phi >= -(180-phi_min) and phi <= -(180-phi_max)) or (phi >= phi_min and phi <= phi_max):
                if phi < 0: phi = 180 + phi
            else:
                return 0
        else:
            if (phi >= -phi_max and phi <= -phi_min) or (phi >= (180-phi_max) and phi <= (180-phi_min)):
                if phi <= 0: phi = abs(phi)
                else: phi = 180 - phi
            else:
                return 0
        return phi

class Json():
    def create_json_from_txt(seft):
        for file_list_test in os.listdir(f'data/list/test_split'):
            file_list_test = file_list_test[:-4]

            #if file_list_test != 'test5_arrow': continue
            
            data_list = []; frame_counter = 0
            if not os.path.exists(f'dataset_phan_loai/{file_list_test}'):
                os.makedirs(f'dataset_phan_loai/{file_list_test}')
            with open(f"data/list/test_split/{file_list_test}.txt", "r") as file:
                lines = file.readlines()
            for line in lines:
                raw_file_old = line.strip()
                raw_file_new = f'dataset_phan_loai/{file_list_test}/{frame_counter:04d}.jpg'
                frame_counter += 1

                #if frame_counter != 135 + 1: continue
                file_txt = f"dataset_goc/{raw_file_old[:-4]}.lines.txt"
                if not os.path.exists(file_txt): 
                    print(f'{file_txt} -   not exits')

                with open(file_txt, 'r') as file:
                    content = file.readlines()
                lanes = []
                for lane in content:
                    lane = lane.strip(); lane = lane.split(); k = 0; arr_xy = []
                    while k < len(lane):
                        x = round(float(lane[k])); y = lane[k+1]; k = k + 2
                        arr_xy.append((int(x), int(y)))
                    lanes.append(arr_xy)

                for i in range(len(lanes)):
                    l = lanes[i]
                    l = [(x,y) for x, y in l if x > 0]
                    l = sorted(l, key=lambda x: x[1], reverse=False)
                    lanes[i] = l

                lane_left, lane_right = [], []
                for l in lanes:
                    l = [[x,y] for x,y in l if y <= y_stop] 
                    if len(l) == 0: continue
                    x_top, y_top = l[0]
                    x_bot, y_bot = l[-1]

                    if x_top - x_bot > 0:
                        if x_bot < (x_mid - 30):
                            lane_left.append(l)
                    elif x_bot - x_top > 0:
                        if x_bot > (x_mid + 30):
                            lane_right.append(l)
                multi_lane_left = lane_left
                multi_lane_right = lane_right

                mid_lane_left, mid_lane_right = [], []

                if len(lane_left) >= 2:
                    for k in range(len(lane_left)-1):
                        l1 = lane_left[k]
                        l2 = lane_left[k+1]
                        x1_bot, y1_top = l1[-1]; x2_bot, y2_top = l2[-1]
                        if x2_bot > x1_bot:
                            mid_lane_left = l2
                        else:
                            mid_lane_left = l1
                elif len(lane_left) == 1:
                    mid_lane_left = lane_left[0]

                if len(lane_right) >= 2:
                    for k in range(len(lane_right)-1):
                        l1 = lane_right[k]
                        l2 = lane_right[k+1]
                        x1_bot, y1_top = l1[-1]; x2_bot, y2_top = l2[-1]
                        if x1_bot < x2_bot:
                            mid_lane_right = l1
                        else:
                            mid_lane_right = l2
                elif len(lane_right) == 1:
                    mid_lane_right = lane_right[0]

                data = {
                    "raw_file_old": f'dataset_goc/{raw_file_old}',
                    "raw_file_new": raw_file_new,
                    "lanes": lanes,
                    "mid_lane_left": mid_lane_left,
                    "mid_lane_right": mid_lane_right,
                    "multi_lane_left": multi_lane_left,
                    "multi_lane_right": multi_lane_right
                }
                data_list.append(data)

                print(f'{file_list_test}  -  {raw_file_old} - {raw_file_new}')

            with open(f"json/true/{file_list_test}.json", "w") as json_file:
                json.dump(data_list, json_file)
    def get_no_lane_of_gt(seft, name):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]; total_frame = len(json_gt); 
        data_list = []; data_list_02 = []

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            mid_lane_left = gt['mid_lane_left']; mid_lane_right = gt['mid_lane_right']

            if len(mid_lane_left) == 0:
                data = {
                    "raw_file_old": raw_file_old,
                    "raw_file_new": raw_file_new,
                }
                data_list.append(data)
            
            if len(mid_lane_right) == 0:
                data = {
                    "raw_file_old": raw_file_old,
                    "raw_file_new": raw_file_new,
                }
                data_list_02.append(data)

        with open(f"json/true/{name}_no_lane_left.json", "w") as json_file:
            json.dump(data_list, json_file)
        with open(f"json/true/{name}_no_lane_right.json", "w") as json_file:
            json.dump(data_list_02, json_file)

class Excel():
    def file_excel_ground_true_said_no_lane(seft, name):
        ws.cell(row=1, column=1).value = "stt"
        ws.cell(row=1, column=2).value = "raw_file_old"
        ws.cell(row=1, column=3).value = "raw_file_new"
        row = 2

        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            mid_lane_left = gt['mid_lane_left']; mid_lane_right = gt['mid_lane_right']; lanes = gt['lanes']

            if len(mid_lane_right) == 0:
                ws.cell(row=row, column=1).value = f'{i:04d}'
                ws.cell(row=row, column=2).value = raw_file_old
                ws.cell(row=row, column=3).value = raw_file_new
                ws.cell(row=row, column=4).value = "không có lane nghiêng bên trái"
                img = cv2.imread(raw_file_old)
                for lane in lanes:
                    for pt in lane:
                        x,y = pt[0], pt[1]
                        cv2.circle(img, (x,y), radius=15, color=(0, 255, ), thickness=1)
                cv2.imwrite(f'my_frame_processing/dont_case/right_{raw_file_new[-8:-4]}.jpg', img)
                row += 1

        #wb.save(f'Analysis/Frame_NO_LINE/new/{name}_không_có_lane_nghiêng_bên_trái.xlsx')
    def file_excel_ratio_acc(seft, name, number_frame):
        ws.cell(row=1, column=1).value = "raw_file_old"
        ws.cell(row=1, column=2).value = "raw_file_new"
        
        ws.cell(row=1, column=3).value = "stt"
        ws.cell(row=1, column=4).value = "L.ratio"
        ws.cell(row=1, column=5).value = "L.acc"
        ws.cell(row=1, column=6).value = "L.Khong co lane"

        ws.cell(row=1, column=7).value = "stt"
        ws.cell(row=1, column=8).value = "R.ratio"
        ws.cell(row=1, column=9).value = "R.acc"
        ws.cell(row=1, column=10).value = "R.Khong co lane"
        row = 2

        json_gt = [json.loads(line) for line in open(f"json/2lane/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}].json")][0]; total_frame = len(json_gt); data_list = []

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']

            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            ws.cell(row=row, column=1).value = raw_file_old
            ws.cell(row=row, column=2).value = raw_file_new

            ws.cell(row=row, column=3).value = f'{raw_file_new[-8:-4]}'
            ws.cell(row=row, column=4).value = f'{ratio_left}'
            ws.cell(row=row, column=5).value = acc_left / 100.

            ws.cell(row=row, column=7).value = f'{raw_file_new[-8:-4]}'
            ws.cell(row=row, column=8).value = f'{ratio_right}'
            ws.cell(row=row, column=9).value = acc_right / 100.

            row += 1

        wb.save(f'Analysis/Frame_NO_LINE/new/{name}_ratio_acc_Frame[{number_frame[0]},{number_frame[1]}].xlsx')
    def get_json_frame_no_lane_from_excel(seft, name):
        df = pd.read_excel(f'Analysis/Frame_NO_LINE/{name}_ratio_acc_Frame[{number_frame[0]},{number_frame[1]}].xlsx')
        str_no_add_acc = ['không có lane', 'gt bị lệch', 'cong', 'mở rộng roi']

        data_list = []
        for tmp in str_no_add_acc:
            filtered_rows = df[df['L.Khong co lane'] == tmp]
            raw_file_new = filtered_rows['raw_file_new'].tolist(); raw_file_old = filtered_rows['raw_file_old'].tolist()
            for i in range(len(raw_file_new)):
                data = {
                    "raw_file_new": raw_file_new[i],
                    "raw_file_old": raw_file_old[i]
                }
                data_list.append(data); print(f'{raw_file_new[i]} - {tmp}')

        with open(f"json/no_add_acc/{name}_left_Frame[{number_frame[0]},{number_frame[1]}].json", "w") as json_file:
            json.dump(data_list, json_file)
        #========================================================================================================================================
        data_list = []
        for tmp in str_no_add_acc:
            filtered_rows = df[df['R.Khong co lane'] == tmp]
            raw_file_new = filtered_rows['raw_file_new'].tolist(); raw_file_old = filtered_rows['raw_file_old'].tolist()
            for i in range(len(raw_file_new)):
                data = {
                    "raw_file_new": raw_file_new[i],
                    "raw_file_old": raw_file_old[i]
                }
                data_list.append(data); print(f'{raw_file_new[i]} - {tmp}')

        with open(f"json/no_add_acc/{name}_right_Frame[{number_frame[0]},{number_frame[1]}].json", "w") as json_file:
            json.dump(data_list, json_file)
    def get_frame_lane_blurred_from_excel(seft, name):
        df = pd.read_excel(f'Analysis/Frame_NO_LINE/{name}_đã_check_frame_0_đến_1000.xlsx')

        filtered_rows = df[df['L.Khong co lane'] == 'lane blurred']
        raw_file_new = filtered_rows['raw_file_new'].tolist()
        raw_file_old = filtered_rows['raw_file_old'].tolist()

        data_list = []
        for i in range(len(raw_file_new)):
            data = {
                "raw_file_new": raw_file_new[i],
                "raw_file_old": raw_file_old[i]
            }
            data_list.append(data)
            print(f'{raw_file_new[i]} - lane blurred left')
        with open(f"json/true/{name}_blurred_lane_left.json", "w") as json_file:
            json.dump(data_list, json_file)

        filtered_rows = df[df['R.Khong co lane'] == 'lane blurred']
        raw_file_new = filtered_rows['raw_file_new'].tolist()
        raw_file_old = filtered_rows['raw_file_old'].tolist()

        data_list = []
        for i in range(len(raw_file_new)):
            data = {
                "raw_file_new": raw_file_new[i],
                "raw_file_old": raw_file_old[i],
            }
            data_list.append(data)
            print(f'{raw_file_new[i]} - lane blurred right')
        with open(f"json/true/{name}_blurred_lane_right.json", "w") as json_file:
            json.dump(data_list, json_file)
    def save_name_raw_file_to_excel(name):
        json_gt = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        total_frame = len(json_gt); data_list = []
        row = 2
        ws.cell(row=1, column=1).value = "raw_file_old"
        ws.cell(row=1, column=2).value = "raw_file_new"
        ws.cell(row=1, column=3).value = "STT"
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            mid_lane_left = gt['mid_lane_left']; mid_lane_right = gt['mid_lane_right']

            ws.cell(row=row, column=1).value = raw_file_old
            ws.cell(row=row, column=2).value = raw_file_new
            ws.cell(row=row, column=3).value = f'{i:04d}'
            row += 1
        wb.save(f'Analysis/Excel/{name}_name_raw_file.xlsx')

class Analysis():
    def get_total_frame(seft):
        arr_total_frame = []
        for file_list_test in os.listdir(f'data/list/test_split'):
            file_list_test = file_list_test[:-4]
            #if file_list_test != 'test8_night': continue
            json_gt = [json.loads(line) for line in open(f'json/true/{file_list_test}.json')][0]; total_frame = len(json_gt)
            arr_total_frame.append((f'{file_list_test}', int(total_frame)))

        for l in arr_total_frame:
            print(f'{l[0]}: {l[1]} frame')
        #=====================================================================================================
        
        json_no_lane_left = [json.loads(line) for line in open(f'json/true/{name}_no_lane_left.json').readlines()][0]
        json_no_lane_right = [json.loads(line) for line in open(f'json/true/{name}_no_lane_right.json').readlines()][0]

        t_no_lane_left = len(json_no_lane_left)
        t_no_lane_right = len(json_no_lane_right)

        json_blurred_lane_left = [json.loads(line) for line in open(f'json/true/{name}_blurred_lane_left.json').readlines()][0]
        json_blurred_lane_right = [json.loads(line) for line in open(f'json/true/{name}_blurred_lane_right.json').readlines()][0]    

        t_blurred_lane_left = len(json_blurred_lane_left)
        t_blurred_lane_right = len(json_blurred_lane_right)
    def get_img_incorrect(seft, name, acc_muc_tieu, number_frame, flag_left):
        json_gt = [json.loads(line) for line in open(f"json/2lane/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}]_them_median.json")][0]
        json_gt_02 = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        gts_02 = {l['raw_file_new']: l for l in json_gt_02}
        total_frame = len(json_gt); data_list = []

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']
            pred_lane_left = gt['pred_lane_left']; pred_lane_right = gt['pred_lane_right']; 

            gt_02 = gts_02[raw_file_new]
            mid_lane_left = gt_02['mid_lane_left']; mid_lane_right = gt_02['mid_lane_right']; 
            multi_lane_left = gt_02['multi_lane_left']; multi_lane_right = gt_02['multi_lane_right']; 
            
            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            if flag_left: acc = acc_left
            else: acc = acc_right

            if acc < acc_muc_tieu:
                img = cv2.imread(raw_file_old)

                cv2.polylines(img, [roi_left_draw], isClosed=True, color=(127,255,0), thickness=2)
                cv2.polylines(img, [roi_right_draw], isClosed=True, color=(255,127,0), thickness=2)

                angles_left = LaneEval.get_angle(mid_lane_left)
                angles_right = LaneEval.get_angle(mid_lane_right)

                left_thresh = LaneEval.pixel_thresh / np.cos(angles_left)
                right_thresh = LaneEval.pixel_thresh / np.cos(angles_right)

                for l in multi_lane_left:
                    for pt in l:
                        x,y = pt[0], pt[1]
                        cv2.circle(img, (x,y), radius=3, color=(255, 127, ), thickness=1)
                        #cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)
                for l in multi_lane_right:
                    for pt in l:
                        x,y = pt[0], pt[1]
                        cv2.circle(img, (x,y), radius=3, color=(255, 127, ), thickness=1)
                        #cv2.line(img, (round(x-right_thresh), y), (round(x+right_thresh), y), (127,127,0), 1)

                for pt in mid_lane_left:
                    x,y = pt[0], pt[1]
                    cv2.circle(img, (x,y), radius=2, color=(0, 255, ), thickness=-1)
                    cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)
                for pt in mid_lane_right:
                    x,y = pt[0], pt[1]
                    cv2.circle(img, (x,y), radius=2, color=(0, 255, ), thickness=-1)
                    cv2.line(img, (round(x-right_thresh), y), (round(x+right_thresh), y), (127,127,0), 1)

                for pt in pred_lane_left:
                    cv2.circle(img, pt, radius=4, color=(0, 0, 255), thickness=1)
                for pt in pred_lane_right:
                    cv2.circle(img, pt, radius=4, color=(0, 0, 255), thickness=1)

                if flag_left:
                    img = img[y_start-10:y_stop+10,x_mid-x_ext_bot-10:x_mid+10,]
                    img = cv2.resize(img,(370*3, 130*3))
                    cv2.imwrite(f'my_frame_processing/img_incorrect/{name}/Frame[{number_frame[0]},{number_frame[1]}]/left/{raw_file_new[-8:-4]}_them_median.jpg', img)
                else:
                    img = img[y_start-10:y_stop+10,x_mid-10:x_mid+x_ext_bot+10,]; 
                    img = cv2.resize(img,(370*3, 130*3))
                    cv2.imwrite(f'my_frame_processing/img_incorrect/{name}/Frame[{number_frame[0]},{number_frame[1]}]/right/{raw_file_new[-8:-4]}_them_median.jpg', img)

                print(f'{raw_file_new}')
    def get_img_incorrect_no_add_no_lane(seft, name, acc_muc_tieu, number_frame, flag_left):
        json_gt = [json.loads(line) for line in open(f"json/2lane/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}].json")][0]
        json_gt_02 = [json.loads(line) for line in open(f'json/true/{name}.json')][0]
        gts_02 = {l['raw_file_new']: l for l in json_gt_02}
        total_frame = len(json_gt); data_list = []

        json_no_lane_left = [json.loads(line) for line in open(f'json/true/{name}_no_lane_left_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]
        json_no_lane_right = [json.loads(line) for line in open(f'json/true/{name}_no_lane_right_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]

        json_no_lane_left = {l['raw_file_new']: l for l in json_no_lane_left}
        json_no_lane_right = {l['raw_file_new']: l for l in json_no_lane_right}

        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']
            pred_lane_left = gt['pred_lane_left']; pred_lane_right = gt['pred_lane_right']; 

            gt_02 = gts_02[raw_file_new]
            mid_lane_left = gt_02['mid_lane_left']; mid_lane_right = gt_02['mid_lane_right']; 
            multi_lane_left = gt_02['multi_lane_left']; multi_lane_right = gt_02['multi_lane_right']; 
            
            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            if flag_left: 
                acc = acc_left
                json_no_lane = json_no_lane_left
            else: 
                acc = acc_right
                json_no_lane = json_no_lane_right

            if raw_file_new in json_no_lane:
                print(f'{raw_file_new} no lane')
            else:
                if acc < acc_muc_tieu:
                    img = cv2.imread(raw_file_old)

                    cv2.polylines(img, [roi_left_draw], isClosed=True, color=(127,255,0), thickness=2)
                    cv2.polylines(img, [roi_right_draw], isClosed=True, color=(255,127,0), thickness=2)

                    angles_left = LaneEval.get_angle(mid_lane_left)
                    angles_right = LaneEval.get_angle(mid_lane_right)

                    left_thresh = LaneEval.pixel_thresh / np.cos(angles_left)
                    right_thresh = LaneEval.pixel_thresh / np.cos(angles_right)

                    for l in multi_lane_left:
                        for pt in l:
                            x,y = pt[0], pt[1]
                            cv2.circle(img, (x,y), radius=3, color=(255, 127, ), thickness=1)
                            #cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)
                    for l in multi_lane_right:
                        for pt in l:
                            x,y = pt[0], pt[1]
                            cv2.circle(img, (x,y), radius=3, color=(255, 127, ), thickness=1)
                            #cv2.line(img, (round(x-right_thresh), y), (round(x+right_thresh), y), (127,127,0), 1)

                    for pt in mid_lane_left:
                        x,y = pt[0], pt[1]
                        cv2.circle(img, (x,y), radius=2, color=(0, 255, ), thickness=-1)
                        cv2.line(img, (round(x-left_thresh), y), (round(x+left_thresh), y), (127,127,0), 1)
                    for pt in mid_lane_right:
                        x,y = pt[0], pt[1]
                        cv2.circle(img, (x,y), radius=2, color=(0, 255, ), thickness=-1)
                        cv2.line(img, (round(x-right_thresh), y), (round(x+right_thresh), y), (127,127,0), 1)

                    for pt in pred_lane_left:
                        cv2.circle(img, pt, radius=4, color=(0, 0, 255), thickness=1)
                    for pt in pred_lane_right:
                        cv2.circle(img, pt, radius=4, color=(0, 0, 255), thickness=1)

                    if flag_left:
                        img = img[y_start-10:y_stop+10,x_mid-x_ext_bot-10:x_mid+10,]
                        img = cv2.resize(img,((x_ext_bot+20)*3, (y_stop-y_start+20)*3))
                        cv2.imwrite(f'my_frame_processing/img_incorrect/{name}/Frame[{number_frame[0]},{number_frame[1]}]/left/{raw_file_new[-8:-4]}.jpg', img)
                    else:
                        img = img[y_start-10:y_stop+10,x_mid-10:x_mid+x_ext_bot+10,]; 
                        img = cv2.resize(img,((x_ext_bot+20)*3, (y_stop-y_start+20)*3))
                        cv2.imwrite(f'my_frame_processing/img_incorrect/{name}/Frame[{number_frame[0]},{number_frame[1]}]/right/{raw_file_new[-8:-4]}.jpg', img)

                    print(f'{raw_file_new}')
    def get_acc_remove_frame_no_bullerd_lane(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f"json/2lane/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}].json")][0]
        total_frame = len(json_gt); data_list = []

        gts = {l['raw_file_new']: l for l in json_gt}
        json_no_lane_left = [json.loads(line) for line in open(f'json/true/{name}_no_lane_left.json').readlines()][0]
        json_no_lane_right = [json.loads(line) for line in open(f'json/true/{name}_no_lane_right.json').readlines()][0]

        json_no_lane_left = {l['raw_file_new']: l for l in json_no_lane_left}
        json_no_lane_right = {l['raw_file_new']: l for l in json_no_lane_right}

        json_blurred_lane_left = [json.loads(line) for line in open(f'json/true/{name}_blurred_lane_left.json').readlines()][0]
        json_blurred_lane_right = [json.loads(line) for line in open(f'json/true/{name}_blurred_lane_right.json').readlines()][0]

        json_blurred_lane_left = {l['raw_file_new']: l for l in json_blurred_lane_left}
        json_blurred_lane_right = {l['raw_file_new']: l for l in json_blurred_lane_right}

        arr_acc_left = []; arr_acc_right = []
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']

            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            if raw_file_new in json_no_lane_left or raw_file_new in json_blurred_lane_left :
                print(f'{raw_file_new} no lane or blurred lane left')
            else:
                arr_acc_left.append(acc_left)
            
            if raw_file_new in json_no_lane_right or raw_file_new in json_blurred_lane_right:
                print(f'{raw_file_new} no lane or blurred lane right')
            else:
                arr_acc_right.append(acc_right)

            #print(f'{raw_file_new} - left: {ratio_left} - {round(acc_left, 2)}, right: {ratio_right} - {round(acc_right, 2)}')  
            
        arr_acc_left = np.mean(arr_acc_left); arr_acc_right = np.mean(arr_acc_right); acc = (arr_acc_left + arr_acc_right) / 2.
        print(f'Frame[{number_frame[0]},{number_frame[1]}] - Mean_acc(Left-Right-avg): {round(arr_acc_left, 2)}%-{round(arr_acc_right, 2)}%-{round(acc, 2)}%')  
    def get_acc_remove_no_lane(seft, name, number_frame):
        json_gt = [json.loads(line) for line in open(f"json/Acc/{name}_Frame[{number_frame[0]},{number_frame[1]}].json")][0]

        gts = {l['raw_file_new']: l for l in json_gt}
        json_no_lane_left = [json.loads(line) for line in open(f'json/no_add_acc/{name}_left_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]
        json_no_lane_right = [json.loads(line) for line in open(f'json/no_add_acc/{name}_right_Frame[{number_frame[0]},{number_frame[1]}].json').readlines()][0]

        json_no_lane_left = {l['raw_file_new']: l for l in json_no_lane_left}
        json_no_lane_right = {l['raw_file_new']: l for l in json_no_lane_right}

        arr_acc_left = []; arr_acc_right = []
        for i in range(0, len(json_gt)):
            gt = json_gt[i]; start_time = time.time(); raw_file_new = gt['raw_file_new']; raw_file_old = gt['raw_file_old']
            acc_left = gt['acc_left']; ratio_left = gt['ratio_left']
            acc_right = gt['acc_right']; ratio_right = gt['ratio_right']

            if int(raw_file_new[-8:-4]) < number_frame[0] or int(raw_file_new[-8:-4]) > number_frame[1]: continue

            if raw_file_new in json_no_lane_left:
                print(f'{raw_file_new} no lane left')
            else:
                arr_acc_left.append(acc_left)
            
            if raw_file_new in json_no_lane_right:
                print(f'{raw_file_new} no lane right')
            else:
                arr_acc_right.append(acc_right)

            print(f'{raw_file_new} - left: {ratio_left} - {round(acc_left, 2)}, right: {ratio_right} - {round(acc_right, 2)}')  
        
        print(f'So lane danh gia Left: {len(arr_acc_left)} - Right: {len(arr_acc_left)}')
        arr_acc_left = np.mean(arr_acc_left); arr_acc_right = np.mean(arr_acc_right); acc = (arr_acc_left + arr_acc_right) / 2.
        print(f'{name}_Frame[{number_frame[0]},{number_frame[1]}] - Mean_acc(Left-Right-avg): {round(arr_acc_left, 2)}%-{round(arr_acc_right, 2)}%-{round(acc, 2)}%')  
   
if __name__ == '__main__':
    arr_name = ['test0_normal', 'test1_crowd', 'test2_hlight', 'test3_shadow', 'test4_noline', 'test5_arrow', 'test6_curve', 'test7_cross', 'test8_night']
    arr_number_frame = [(0,0)]
    name = 'test0_normal'
    

    #=====Class Save_csv()======================================================================================================================
    number_frame = 180
    name = arr_name[8]
    # Save_csv().Sobel_get_G(name, number_frame)
    # Save_csv().Median5x5_Sobel_get_G(name, number_frame)
    # Save_csv().Median3x3_Sobel_get_G(name, number_frame)
    #Save_csv().Gaussian3x3_Sobel_get_G(name, number_frame)
    # Save_csv().Gaussian5x5_Sobel_get_G(name, number_frame)
    # Save_csv().Grayscale(name, number_frame)
    # Save_csv().Grayscale_Median5x5(name, number_frame)
    #Save_csv().Grayscale_Gaussian3x3(name, number_frame)
    # Save_csv().Grayscale_Gaussian5x5(name, number_frame)

    #=========Class Draw()================================================================================================================
    name = 'test2_hlight'; number_frame = [0,485]
    #name = 'test0_normal'; number_frame = [7000,9620]
    #Draw().Draw_Roi_and_Lane_True(name, number_frame=number_frame)
    #Draw().Draw_Only_Roi(name, number_frame)
    #Draw().Connect_frame_into_video_demo(name, number_frame)
    #Draw().Draw_predict_gt_text(name, number_frame)
    #Draw().Draw_frame_censorship(name, number_frame, option='khong_co_lane')

    #======= Class Json()==============================================================================================================
    #Json().create_json_from_txt()

    #======= Class Excel()==============================================================================================================
    #name = 'test0_normal'; number_frame = [7000,9620]
    # name = 'test2_hlight'; number_frame = [0,485]
    # name = 'test3_shadow'; number_frame = [0,929]
    # name = 'test5_arrow'; number_frame = [0,889]
    # name = 'test8_night'; number_frame = [0,7028]
    #Excel().file_excel_ground_true_said_no_lane(name)
    #Excel().file_excel_ratio_acc(name, number_frame)
    Excel().get_json_frame_no_lane_from_excel(name)
    #Excel().save_name_raw_file_to_excel(name)

    #======= Class Analysis()==============================================================================================================
    #name = 'test0_normal'; number_frame = [7000,9620]
    name = 'test2_hlight'; number_frame = [0,485]
    # name = 'test3_shadow'; number_frame = [0,929]
    # name = 'test5_arrow'; number_frame = [0,889]
    # name = 'test8_night'; number_frame = [0,7028]
    #Analysis().get_total_frame(name)
    #Analysis().get_img_incorrect(name, acc_muc_tieu = 90, number_frame = number_frame, flag_left=0)
    #Analysis().get_img_incorrect_no_add_no_lane(name, acc_muc_tieu = 90, number_frame = number_frame, flag_left=1)
    Analysis().get_acc_remove_no_lane(name, number_frame)
    #Analysis().get_frame_acc_greater_90(name='test8_night', acc_muc_tieu=90, number_frame=[0,7028])


    print(f'=====================DONE module_bo_tro==================')