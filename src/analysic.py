import json
from lane_lib_TU import LaneEval
import numpy as np
import openpyxl
import cv2
import matplotlib.pyplot as plt

folder_Analysis = "Analysis"

name = "0531"
gt_file = f'json/true_{name}.json'
pred_file = f'json/predict_{name}.json'

write_excel = 1
# pixel_thresh = 20
# pt_thresh = 0.85


def main():
    total_correct = 0

    wb = openpyxl.Workbook()
    ws = wb.active
    row = 2
    col = 1
    ws.cell(row=1, column=col).value = "name"
    ws.cell(row=1, column=col+1).value = "left correct"
    ws.cell(row=1, column=col+2).value = "left incorrect"
    ws.cell(row=1, column=col+3).value = "left thresh"

    # ws.cell(row=1, column=col+3).value = "correct_right"
    # ws.cell(row=1, column=col+4).value = "incorrect_right"
    
    #ws.cell(row=1, column=col+6).value = "threshs_right"

    try:
        json_pred = [json.loads(line) for line in open(pred_file).readlines()]
    except BaseException as e:
        raise Exception('Fail to load json file of the prediction.')
    json_gt = [json.loads(line) for line in open(gt_file).readlines()]

    json_gt = json_gt[0]
    json_pred = json_pred[0]

    data_list = []
    data_list_left = []
    data_list_right = []
    y_samples = json_gt[0]['h_samples']
    len_y_samples = len(y_samples)

    # if len(json_gt) != len(json_pred):
    #     raise Exception('We do not get the predictions of all the test tasks')
    gts = {l['raw_file']: l for l in json_gt}
    accuracy, fp, fn = 0., 0., 0.

    for pred in json_pred:
        raw_file = pred['raw_file']
        pred_lanes = pred['lanes']
        if raw_file not in gts:
            raise Exception('Some raw_file from your predictions do not exist in the test tasks.')
        gt = gts[raw_file]
        gt_lanes = gt['lanes']
        
        
        angles_left = LaneEval.get_angle(np.array(gt_lanes[0]), np.array(y_samples))
        angles_right = LaneEval.get_angle(np.array(gt_lanes[1]), np.array(y_samples))

        threshs_left = LaneEval.pixel_thresh / np.cos(angles_left)
        threshs_right = LaneEval.pixel_thresh / np.cos(angles_right)

        sum_point_correct_left = LaneEval.sum_point_correct(pred_lanes[0], gt_lanes[0], threshs_left)
        sum_point_incorrect_left = len_y_samples - sum_point_correct_left
        total_correct += sum_point_correct_left

        sum_point_correct_right = LaneEval.sum_point_correct(pred_lanes[1], gt_lanes[1], threshs_right)
        sum_point_incorrect_right = len_y_samples - sum_point_correct_right
        total_correct += sum_point_correct_right

        # if sum_point_correct_left < 48:
        #     img = cv2.imread(raw_file)
        #     img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        #     draw_predi_left = [[(x, y) for (x, y) in zip(pred_lanes[0], y_samples) if x >= 0] ]

        #     for lane in draw_predi_left:
        #         cv2.polylines(img, np.int32([lane]), isClosed=False, color=(0,0,255), thickness=2)

        #     draw_gt_lane = [[(x, y) for (x, y) in zip(lane, y_samples) if x >= 0] for lane in gt_lanes]
        #     for lane in draw_gt_lane:
        #         for pt in lane:
        #             cv2.circle(img, pt, radius=5, color=(0, 255, 0), thickness=2)
        #     for height in y_samples:
        #         start_point = (0, height)  # Starting from the left, at the specified height
        #         end_point = (img.shape[1], height)  # Ending at the right, at the specified height
        #         cv2.line(img, start_point, end_point, (255, 0, 0), 1)

        #     plt.imshow(img)
        #     plt.title(raw_file[-15:-9])
        #     plt.show()
        #     print("stop")

        # if sum_point_correct_right < 48:
        #     list_wrong = []
        #     denta_arr = np.array(gt_lanes[1]) - np.array(pred_lanes[1])
        #     for i in range(0, len_y_samples):
        #         denta = round(abs(denta_arr[i]))
        #         if denta > threshs_right:
        #             data = (y_samples[i], denta)
        #             list_wrong.append(data)
        #     print(list_wrong)
            
        #     img = cv2.imread(raw_file)
        #     img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        #     draw_predi_right = [[(x, y) for (x, y) in zip(pred_lanes[1], y_samples) if x >= 0] ]

        #     for lane in draw_predi_right:
        #         cv2.polylines(img, np.int32([lane]), isClosed=False, color=(0,0,255), thickness=2)

        #     draw_gt_lane = [[(x, y) for (x, y) in zip(lane, y_samples) if x >= 0] for lane in gt_lanes]
        #     for lane in draw_gt_lane:
        #         for pt in lane:
        #             cv2.circle(img, pt, radius=5, color=(0, 255, 0), thickness=2)
        #     for height in y_samples:
        #         start_point = (0, height)  # Starting from the left, at the specified height
        #         end_point = (img.shape[1], height)  # Ending at the right, at the specified height
        #         cv2.line(img, start_point, end_point, (255, 0, 0), 1)

        #     plt.imshow(img)
        #     plt.title(raw_file[-15:-9])
        #     plt.show()
            



            # ws.cell(row=row, column=col).value = raw_file[-15:-9]
            # ws.cell(row=row, column=col+1).value = sum_point_correct_left
            # ws.cell(row=row, column=col+2).value = sum_point_incorrect_left
            # ws.cell(row=row, column=col+3).value = threshs_left

            # ws.cell(row=row, column=col+4).value = "y_samples"
            # ws.cell(row=row+1, column=col+4).value = "gt_lane[0]"
            # ws.cell(row=row+2, column=col+4).value = "pred_lane[0]"
            # ws.cell(row=row+3, column=col+4).value = "gt - pred"
            
            # for i, value in enumerate(y_samples):
            #     ws.cell(row=row, column=i+6).value = value
            # row += 1
            # for i, value in enumerate(gt_lanes[0]):
            #     ws.cell(row=row, column=i+6).value = value
            # row += 1
            # for i, value in enumerate(pred_lanes[0]):
            #     ws.cell(row=row, column=i+6).value = value
            # row += 1
            # denta = np.array(gt_lanes[0]) - np.array(pred_lanes[0])
            # for i, value in enumerate(denta):
            #     ws.cell(row=row, column=i+6).value = value   
            # row += 1

        #if sum_point_correct_right < 48:
            # data = {
            #     "raw_file": raw_file,
            #     "gt_lane_right": gt_lanes[1],
            #     "pred_lanes_right": pred_lanes[1],
            # }
            # data_list_right.append(data)
        # data = {
        #     "raw_file": raw_file,
        #     "sum_point_correct_left": sum_point_correct_left,
        #     "sum_point_incorrect_left": sum_point_incorrect_left,
        #     "sum_point_correct_right": sum_point_correct_right,
        #     "sum_point_incorrect_right": sum_point_incorrect_right,
        #     "gt_lane": gt_lanes,
        #     "pred_lanes": pred_lanes,
        # }

        # data_list.append(data)

        # ws.cell(row=row, column=col).value = raw_file[-15:-9]
        # ws.cell(row=row, column=col+1).value = sum_point_correct_left
        # ws.cell(row=row, column=col+2).value = sum_point_incorrect_left

        # ws.cell(row=row, column=col+3).value = sum_point_correct_right
        # ws.cell(row=row, column=col+4).value = sum_point_incorrect_right
        # ws.cell(row=row, column=col+5).value = threshs_left
        # ws.cell(row=row, column=col+6).value = threshs_right
            # row += 1
    if write_excel:  
        wb.save(folder_Analysis + '/false_lane_left.xlsx')
    # accuracy += a
    # fp += p
    # fn += n

    # num = len(gts)
    # Accuracy  = accuracy / num
    # FP = fp / num
    # FN = fn / num
    # print('Accuracy: ', Accuracy)
    # print('FP: ', FP)
    # print('FN: ', FN)
    accuracy = total_correct / (len(json_pred) * len_y_samples * 2)
    return accuracy
if __name__ == '__main__':
    accuracy = main()
    print('accuracy: ', accuracy)
    print("===done analysic===")