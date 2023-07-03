
import json
import numpy as np
import matplotlib.pyplot as plt
from lane_lib_TU import LaneEval
import math
import video_lib as Vlib
import openpyxl



folder_Analysis = 'Analysis'



def get_phi_of_TU():
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 1
    col = 1
    phi_arr_left = []
    phi_arr_right = []
    for i in range(0, len(json_gt)):
        gt = json_gt[i]
        gt_lanes = gt['lanes']
        raw_file = gt['raw_file']
        y_samples = gt['h_samples']
        
        ws.cell(row=row, column=col).value = raw_file[-15:-9]
        angles = [LaneEval.get_angle(np.array(x_gts), np.array(y_samples)) for x_gts in gt_lanes]
        phi_degree_left = round(math.degrees(angles[0])) + 90
        phi_degree_right = round(math.degrees(angles[1])) + 90
        if phi_degree_left > 90:
            phi_degree_left = 90
        if phi_degree_right < 90:
            phi_degree_right = 90
        if phi_degree_right > 180:
            phi_degree_right = 180
        ws.cell(row=row, column=col+1).value = phi_degree_left
        phi_arr_left.append(phi_degree_left)
        ws.cell(row=row, column=col+2).value = phi_degree_right
        phi_arr_right.append(phi_degree_right)
        row += 1
        
    wb.save(folder_Analysis + '/phi_true_label_Y320.xlsx')

    mean_phi_left = np.mean(phi_arr_left) 
    mean_phi_right = np.mean(phi_arr_right)
    print('phi_left_min_max = ', np.min(phi_arr_left), np.max(phi_arr_left))
    print('phi_left_avg = ', mean_phi_left)
    print('phi_right_min_max = ', np.min(phi_arr_right), np.max(phi_arr_right))
    print('phi_left_avg = ', mean_phi_right)
    
#     num_bins = len(phi_arr_left)

#     hist, bins, _ = plt.hist(phi_arr_left, bins=num_bins, range=(0, 90), edgecolor='blue', linewidth=2)
# # Display the count/frequency above each bar
#     for i in range(len(hist)):
#         plt.text(bins[i], hist[i], str(int(hist[i])))

#     plt.xlabel('Value')
#     plt.ylabel('Frequency')
#     plt.title('phi_arr_left')
#     plt.show()

#     num_bins = len(phi_arr_right)

#     plt.hist(phi_arr_right, bins=num_bins, range=(90, 180), edgecolor='blue', linewidth=2)

#     plt.xlabel('Value')
#     plt.ylabel('Frequency')
#     plt.title('phi_arr_right')
#     plt.show()


# def get_phi_img():
#     phi_img = np.zeros((720, 1280), dtype=np.int16)
#     for i in range(0, len(json_gt)):
#         gt = json_gt[i]
#         raw_file = gt['raw_file']
#         img = cv2.imread(raw_file, cv2.IMREAD_GRAYSCALE)
#         left_phi_img, left_rho_img = Vlib.get_Rho_phi_left(img[:, 0:640])
#         right_phi_img, right_rho_img = Vlib.get_Rho_phi_right(img[:, 640:1280])

#         phi_img[:, 0:640] = left_phi_img
#         phi_img[:, 640:1280] = right_phi_img

#         path_csv = folder_Analysis + '/' + raw_file[-11:-7] + '.csv'
#         np.savetxt(path_csv, phi_img, delimiter=',')

if __name__ == '__main__':
    json_gt = [json.loads(line) for line in open('json/true_label_Y320.json')]
    json_gt = json_gt[0]

    get_phi_of_TU()


    print("DONE")





