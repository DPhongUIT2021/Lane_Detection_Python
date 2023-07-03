
import json
import numpy as np
import cv2
import openpyxl
import matplotlib.pyplot as plt

name = "label"
K_list_left = []
K_list_right = []

wb = openpyxl.Workbook()
ws = wb.active
row = 1
col = 1

def deftec_Y():
    json_gt = [json.loads(line) for line in open(f'json/true_{name}.json')]
    json_gt = json_gt[0]
    y_samples = [160, 170, 180, 190, 200, 210, 220, 230, 240, 250, 260, 270, 280, 290, 300, 310, 320, 330, 340, 350, 360, 370, 380, 390, 400, 410, 420, 430, 440, 450, 460, 470, 480, 490, 500, 510, 520, 530, 540, 550, 560, 570, 580, 590, 600, 610, 620, 630, 640, 650, 660, 670, 680, 690, 700, 710]
    len_y_sample = len(y_samples)

    for i in range(0, len(json_gt)):
        gt = json_gt[i] 
        raw_file = gt['raw_file']
        gt_lanes = gt['lanes']

        for i in range(1, len_y_sample-1):
            denta_top_left = gt_lanes[0][i] - gt_lanes[0][i-1]
            denta_buttom_left = gt_lanes[0][i] - gt_lanes[0][i+1]
            if denta_top_left > 0 and denta_buttom_left > 0:
                if y_samples[i] < 340:
                    K_list_left.append(y_samples[i])
                    ws.cell(row=row, column=1).value = y_samples[i]
                    row += 1
            denta_top_right = gt_lanes[0][i] - gt_lanes[0][i-1]
            denta_buttom_right = gt_lanes[0][i] - gt_lanes[0][i+1]
            if denta_top_right < 0 and denta_buttom_right < 0:
                if y_samples[i] < 340:
                    K_list_right.append(y_samples[i])
                    ws.cell(row=row, column=1).value = y_samples[i]
                    row += 1

    print(len(K_list_left), len(K_list_right))
    print(K_list_left)
    print(K_list_right)
    wb.save('Analysis/detect_Y.xlsx')
    print("Done detect Y")

def detect_X():
    row = 1
    index_left = 0
    index_right = 0
    json_gt = [json.loads(line) for line in open(f'json/true_label_Y320.json')]
    json_gt = json_gt[0]
    y_samples = [160, 170, 180, 190, 200, 210, 220, 230, 240, 250, 260, 270, 280, 290, 300, 310, 320, 330, 340, 350, 360, 370, 380, 390, 400, 410, 420, 430, 440, 450, 460, 470, 480, 490, 500, 510, 520, 530, 540, 550, 560, 570, 580, 590, 600, 610, 620, 630, 640, 650, 660, 670, 680, 690, 700, 710]
    len_y_sample = len(y_samples)

    for i in range(0, len(json_gt)):
        gt = json_gt[i] 
        gt_lanes = gt['lanes']
        
        # for i in range(len(gt_lanes[0]) - 1, -1, -1):
        #     if gt_lanes[0][i] > 0:
        #         index_left = i
        #         break
            
        # for i in range(len(gt_lanes[1]) - 1, -1, -1):
        #     if gt_lanes[1][i] > 0:
        #         index_right = i
        #         break
        
        ws.cell(row=row, column=1).value = gt_lanes[0][index_left]
        K_list_left.append(gt_lanes[0][index_left])
        ws.cell(row=row, column=2).value = gt_lanes[1][index_right]
        K_list_right.append(gt_lanes[1][index_right])
        row += 1
        

    print(len(K_list_left), len(K_list_right))
    print(K_list_left)
    print(K_list_right)
    wb.save('Analysis/detect_X_Top.xlsx')
    print("Done Detect X_top")

def deftec_Y_with_X0_X1280_Y320():
    row_left = 1
    row_right = 1
    json_gt = [json.loads(line) for line in open(f'json/true_label_Y320.json')]
    json_gt = json_gt[0]
    y_samples = json_gt[0]['h_samples']
    len_y_sample = len(y_samples)

    for i in range(0, len(json_gt)):
        gt = json_gt[i] 
        raw_file = gt['raw_file']
        gt_lanes = gt['lanes']

        img = cv2.imread(raw_file)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        draw_gt_lane = [[(x, y) for (x, y) in zip(lane, y_samples) if x >= 0] for lane in gt_lanes]
        for lane in draw_gt_lane:
            for pt in lane:
                cv2.circle(img, pt, radius=5, color=(0, 255, 0), thickness=2)
        for height in y_samples:
            start_point = (0, height)  # Starting from the left, at the specified height
            end_point = (img.shape[1], height)  # Ending at the right, at the specified height
            cv2.line(img, start_point, end_point, (255, 0, 0), 1)

        # plt.imshow(img)
        # plt.title(raw_file[-15:-9])
        # plt.show()
        
        for i in range(len_y_sample-1,-1,-1):
            if gt_lanes[0][i] > 0 and gt_lanes[0][i] <= 20:
                #K_list_left.append(y_samples[i])
                ws.cell(row=row_left, column=1).value = y_samples[i]
                ws.cell(row=row_left, column=2).value = gt_lanes[0][i]
                row_left += 1
                break
        for i in range(len_y_sample-1,-1,-1):
            if gt_lanes[1][i] > 0 and gt_lanes[1][i] >= 1260 :
                #K_list_right.append(y_samples[i])
                ws.cell(row=row_right, column=4).value = y_samples[i]
                ws.cell(row=row_right, column=5).value = gt_lanes[1][i]
                row_right += 1
                break


    # print(len(K_list_left), len(K_list_right))
    # print(K_list_left)
    # print(K_list_right)
    wb.save('Analysis/detect_Y_with_X0_X1280_Y320.xlsx')
    print("Done detect Y")



if __name__ == '__main__':
    deftec_Y_with_X0_X1280_Y320()



